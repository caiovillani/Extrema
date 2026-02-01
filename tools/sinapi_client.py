"""
Cliente para consulta de precos SINAPI.

O SINAPI (Sistema Nacional de Pesquisa de Custos e Indices da
Construcao Civil) e mantido pela Caixa Economica Federal e IBGE.

Dados disponiveis em:
https://www.caixa.gov.br/poder-publico/modernizacao-gestao/sinapi/

As tabelas SINAPI sao publicadas mensalmente em formato XLS/CSV.
Este cliente busca as planilhas, faz cache local, e permite
consultas por codigo ou descricao.
"""

import csv
import logging
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from tools.http_utils import CachedHTTPClient, HTTPError

logger = logging.getLogger(__name__)

# URL base para download das tabelas SINAPI (Caixa)
SINAPI_DOWNLOAD_BASE = (
    "https://www.caixa.gov.br/Downloads/sinapi-a-partir-jul-2009"
)

# Diretorio para cache local de planilhas baixadas
SINAPI_CACHE_DIR = Path(
    os.environ.get("SINAPI_CACHE_DIR", "data/sinapi")
)


@dataclass
class SINAPIComposicao:
    """Representa uma composicao SINAPI."""

    codigo: str
    descricao: str
    unidade: str
    preco_unitario: float
    referencia_mes: str
    estado: str
    desonerado: bool


@dataclass
class SINAPIInsumo:
    """Representa um insumo SINAPI."""

    codigo: str
    descricao: str
    unidade: str
    preco: float
    origem: str
    referencia_mes: str
    estado: str


class SINAPIClient:
    """Cliente para consultas na base SINAPI.

    O fluxo e:
    1. Baixar planilha mensal da Caixa (ou usar cache local)
    2. Parsear CSV/XLS para extrair composicoes e insumos
    3. Indexar em memoria para consultas rapidas
    """

    def __init__(
        self,
        estado: str = "MG",
        http: Optional[CachedHTTPClient] = None,
    ):
        self.estado = estado
        self.referencia_mes = datetime.now().strftime("%Y-%m")
        self.http = http or CachedHTTPClient()
        self._composicoes: dict[str, SINAPIComposicao] = {}
        self._insumos: dict[str, SINAPIInsumo] = {}
        self._loaded = False

    def load_from_csv(self, csv_path: str | Path):
        """
        Carrega composicoes a partir de um arquivo CSV local.

        O CSV deve ter colunas separadas por ponto-e-virgula com
        headers contendo CODIGO, DESCRICAO, UNIDADE, PRECO.

        Args:
            csv_path: Caminho para o arquivo CSV
        """
        path = Path(csv_path)
        if not path.exists():
            logger.warning("CSV not found: %s", path)
            return

        with path.open(encoding="latin-1") as fh:
            reader = csv.DictReader(fh, delimiter=";")
            for row in reader:
                codigo = (
                    row.get("CODIGO", "")
                    or row.get("CODIGO COMPOSICAO", "")
                    or row.get("CODIGO DA COMPOSICAO", "")
                ).strip()
                if not codigo:
                    continue

                descricao = (
                    row.get("DESCRICAO", "")
                    or row.get("DESCRICAO DA COMPOSICAO", "")
                    or row.get("DESCRICAO DO SERVICO", "")
                ).strip()

                unidade = (
                    row.get("UNIDADE", "")
                    or row.get("UNIDADE DE MEDIDA", "")
                ).strip()

                preco_str = (
                    row.get("PRECO UNITARIO", "")
                    or row.get("CUSTO TOTAL", "")
                    or row.get("PRECO", "")
                ).strip()
                preco_str = (
                    preco_str.replace(".", "").replace(",", ".")
                )
                try:
                    preco = float(preco_str) if preco_str else 0.0
                except ValueError:
                    preco = 0.0

                self._composicoes[codigo] = SINAPIComposicao(
                    codigo=codigo,
                    descricao=descricao,
                    unidade=unidade,
                    preco_unitario=preco,
                    referencia_mes=self.referencia_mes,
                    estado=self.estado,
                    desonerado=False,
                )

        self._loaded = True
        logger.info(
            "Loaded %d composicoes from %s",
            len(self._composicoes), path,
        )

    async def ensure_loaded(self):
        """
        Ensure data is loaded. Tries local cache first, then
        downloads from Caixa website.
        """
        if self._loaded:
            return

        # Try local cache directory
        cache_dir = SINAPI_CACHE_DIR / self.estado.lower()
        if cache_dir.exists():
            csvs = sorted(cache_dir.glob("*.csv"), reverse=True)
            if csvs:
                self.load_from_csv(csvs[0])
                return
            xls_files = sorted(
                cache_dir.glob("*.xls*"), reverse=True
            )
            if xls_files:
                self._load_xls(xls_files[0])
                return

        # Try downloading the latest table
        await self._download_latest()

    async def _download_latest(self):
        """
        Attempt to download the latest SINAPI CSV from Caixa.

        The Caixa publishes files with naming conventions like:
        SINAPI_ref_Preco_Comp_MG_202601_NaoDesonerado.csv
        """
        ref = datetime.now().strftime("%Y%m")
        estado = self.estado.upper()
        filename = (
            f"SINAPI_ref_Preco_Comp_{estado}_{ref}"
            f"_NaoDesonerado.csv"
        )
        url = f"{SINAPI_DOWNLOAD_BASE}/{filename}"

        try:
            raw = await self.http.get_bytes(url)
            cache_dir = SINAPI_CACHE_DIR / estado.lower()
            cache_dir.mkdir(parents=True, exist_ok=True)
            out_path = cache_dir / filename
            out_path.write_bytes(raw)
            self.load_from_csv(out_path)
            logger.info(
                "Downloaded SINAPI table: %s", out_path
            )
        except HTTPError as exc:
            logger.warning(
                "Could not download SINAPI table (%s). "
                "Place CSV files in %s manually.",
                exc, SINAPI_CACHE_DIR,
            )

    def _load_xls(self, path: Path):
        """Attempt to parse an XLS file using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            logger.warning(
                "openpyxl not installed; cannot parse %s", path
            )
            return

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        if ws is None:
            return

        headers = [
            str(c.value or "").upper().strip()
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        code_col = None
        desc_col = None
        unit_col = None
        price_col = None

        for i, h in enumerate(headers):
            if "CODIGO" in h:
                code_col = i
            elif "DESCRI" in h:
                desc_col = i
            elif "UNIDADE" in h or h == "UN":
                unit_col = i
            elif "PRECO" in h or "CUSTO" in h:
                price_col = i

        if code_col is None:
            return

        for row in ws.iter_rows(min_row=2, values_only=True):
            codigo = str(row[code_col] or "").strip()
            if not codigo:
                continue
            descricao = str(
                row[desc_col] if desc_col is not None else ""
            ).strip()
            unidade = str(
                row[unit_col] if unit_col is not None else ""
            ).strip()
            preco_raw = (
                row[price_col]
                if price_col is not None
                else 0
            )
            try:
                preco = float(preco_raw) if preco_raw else 0.0
            except (ValueError, TypeError):
                preco = 0.0

            self._composicoes[codigo] = SINAPIComposicao(
                codigo=codigo,
                descricao=descricao,
                unidade=unidade,
                preco_unitario=preco,
                referencia_mes=self.referencia_mes,
                estado=self.estado,
                desonerado=False,
            )

        self._loaded = True
        logger.info(
            "Loaded %d composicoes from XLS %s",
            len(self._composicoes), path,
        )

    def get_composicao(
        self,
        codigo: str,
        desonerado: bool = False,
    ) -> Optional[SINAPIComposicao]:
        """
        Busca composicao por codigo SINAPI.

        Args:
            codigo: Codigo SINAPI (ex: "87529")
            desonerado: Se True, usa tabela desonerada

        Returns:
            SINAPIComposicao ou None se nao encontrado
        """
        comp = self._composicoes.get(codigo)
        if comp and desonerado != comp.desonerado:
            return None
        return comp

    def search_composicoes(
        self,
        termo: str,
        desonerado: bool = False,
        limite: int = 10,
    ) -> list[SINAPIComposicao]:
        """
        Busca composicoes por descricao.

        Args:
            termo: Termo de busca na descricao
            desonerado: Se True, usa tabela desonerada
            limite: Maximo de resultados

        Returns:
            Lista de SINAPIComposicao
        """
        pattern = re.compile(re.escape(termo), re.IGNORECASE)
        results: list[SINAPIComposicao] = []
        for comp in self._composicoes.values():
            if desonerado != comp.desonerado:
                continue
            if pattern.search(comp.descricao):
                results.append(comp)
                if len(results) >= limite:
                    break
        return results

    def get_insumo(
        self, codigo: str
    ) -> Optional[SINAPIInsumo]:
        """
        Busca insumo por codigo SINAPI.

        Args:
            codigo: Codigo do insumo SINAPI

        Returns:
            SINAPIInsumo ou None
        """
        return self._insumos.get(codigo)

    def calcular_composicao_com_bdi(
        self,
        composicoes: list[tuple[str, float]],
        bdi: float,
    ) -> dict:
        """
        Calcula valor total de composicoes aplicando BDI.

        Args:
            composicoes: Lista de tuplas (codigo, quantidade)
            bdi: Percentual de BDI (ex: 22.12 para 22,12%)

        Returns:
            dict com valores detalhados e total
        """
        items = []
        subtotal = 0.0
        for codigo, qtd in composicoes:
            comp = self._composicoes.get(codigo)
            if comp is None:
                items.append({
                    "codigo": codigo,
                    "descricao": "NAO ENCONTRADO",
                    "quantidade": qtd,
                    "preco_unitario": 0.0,
                    "preco_total": 0.0,
                })
                continue
            total_item = comp.preco_unitario * qtd
            subtotal += total_item
            items.append({
                "codigo": codigo,
                "descricao": comp.descricao,
                "unidade": comp.unidade,
                "quantidade": qtd,
                "preco_unitario": comp.preco_unitario,
                "preco_total": round(total_item, 2),
            })

        bdi_valor = round(subtotal * (bdi / 100), 2)
        return {
            "composicoes": items,
            "subtotal": round(subtotal, 2),
            "bdi_percentual": bdi,
            "bdi_valor": bdi_valor,
            "total": round(subtotal + bdi_valor, 2),
            "referencia": (
                f"SINAPI {self.estado} {self.referencia_mes}"
            ),
        }
