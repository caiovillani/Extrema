"""
Cliente para o Banco de Precos em Saude (BPS) e tabela CMED.

BPS: Mantido pelo Ministerio da Saude, registra precos praticados
em compras publicas de medicamentos e insumos de saude.
URL: https://bps.saude.gov.br/

CMED: Camara de Regulacao do Mercado de Medicamentos (ANVISA).
Publica tabela com precos maximos de medicamentos.
URL: https://www.gov.br/anvisa/pt-br/assuntos/medicamentos/cmed
"""

import csv
import io
import logging
import os
import re
import statistics
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from tools.http_utils import CachedHTTPClient, HTTPError

logger = logging.getLogger(__name__)

# URL do BPS (portal web, nao possui API REST publica)
BPS_URL = "https://bps.saude.gov.br/visao/consultaPublica/index.xhtml"

# URL base da tabela CMED (ANVISA publica em XLS)
CMED_URL = (
    "https://www.gov.br/anvisa/pt-br/assuntos/medicamentos/cmed/"
    "precos"
)

# Diretorio para cache local da tabela CMED
CMED_CACHE_DIR = Path(
    os.environ.get("CMED_CACHE_DIR", "data/cmed")
)


@dataclass
class BPSPreco:
    """Representa um registro de preco no BPS."""

    medicamento: str
    apresentacao: str
    principio_ativo: str
    preco_unitario: float
    orgao_comprador: str
    uf: str
    data_compra: str
    quantidade: int
    modalidade: str


@dataclass
class BPSResumo:
    """Resumo estatistico de precos no BPS."""

    medicamento: str
    apresentacao: str
    media: float
    mediana: float
    minimo: float
    maximo: float
    desvio_padrao: float
    n_registros: int
    periodo: str


@dataclass
class CMEDPreco:
    """Preco maximo de medicamento conforme CMED/ANVISA."""

    medicamento: str
    apresentacao: str
    laboratorio: str
    pf_sem_impostos: float
    pmvg_sem_impostos: float
    pmvg_com_impostos: float
    lista_concessao: str
    data_publicacao: str


class BPSClient:
    """Cliente para consultas no BPS e CMED.

    O BPS nao possui API REST publica; a interacao real requer
    automacao do formulario web. Este cliente fornece:

    1. Parsing de dados BPS previamente exportados (CSV)
    2. Consulta a tabela CMED (download + parsing de XLS/CSV)
    3. Verificacao de teto CMED para medicamentos
    """

    def __init__(
        self, http: Optional[CachedHTTPClient] = None
    ):
        self.http = http or CachedHTTPClient()
        self._cmed_data: dict[str, CMEDPreco] = {}
        self._bps_data: list[BPSPreco] = []
        self._cmed_loaded = False
        self._bps_loaded = False

    # ------------------------------------------------------------------
    # BPS methods
    # ------------------------------------------------------------------

    def load_bps_csv(self, csv_path: str | Path):
        """
        Carrega registros BPS de um arquivo CSV exportado do portal.

        O CSV deve ter colunas: MEDICAMENTO, APRESENTACAO,
        PRINCIPIO_ATIVO, PRECO_UNITARIO, ORGAO, UF, DATA_COMPRA,
        QUANTIDADE, MODALIDADE.

        Args:
            csv_path: Caminho para o CSV exportado
        """
        path = Path(csv_path)
        if not path.exists():
            logger.warning("BPS CSV not found: %s", path)
            return

        with path.open(encoding="utf-8") as fh:
            reader = csv.DictReader(fh, delimiter=";")
            for row in reader:
                preco_str = (
                    row.get("PRECO_UNITARIO", "0")
                    .replace(".", "")
                    .replace(",", ".")
                )
                try:
                    preco = float(preco_str)
                except ValueError:
                    preco = 0.0

                qtd_str = row.get("QUANTIDADE", "0").strip()
                try:
                    qtd = int(qtd_str)
                except ValueError:
                    qtd = 0

                self._bps_data.append(BPSPreco(
                    medicamento=row.get(
                        "MEDICAMENTO", ""
                    ).strip(),
                    apresentacao=row.get(
                        "APRESENTACAO", ""
                    ).strip(),
                    principio_ativo=row.get(
                        "PRINCIPIO_ATIVO", ""
                    ).strip(),
                    preco_unitario=preco,
                    orgao_comprador=row.get(
                        "ORGAO", ""
                    ).strip(),
                    uf=row.get("UF", "").strip(),
                    data_compra=row.get(
                        "DATA_COMPRA", ""
                    ).strip(),
                    quantidade=qtd,
                    modalidade=row.get(
                        "MODALIDADE", ""
                    ).strip(),
                ))

        self._bps_loaded = True
        logger.info(
            "Loaded %d BPS records from %s",
            len(self._bps_data), path,
        )

    def search_bps(
        self,
        medicamento: str,
        apresentacao: Optional[str] = None,
        uf: Optional[str] = None,
        periodo_meses: int = 12,
    ) -> Optional[BPSResumo]:
        """
        Busca precos de medicamento nos dados BPS carregados.

        Args:
            medicamento: Nome ou principio ativo
            apresentacao: Forma farmaceutica e concentracao
            uf: Filtro por UF
            periodo_meses: Periodo de busca em meses

        Returns:
            BPSResumo com estatisticas de precos ou None
        """
        pattern = re.compile(
            re.escape(medicamento), re.IGNORECASE
        )
        matches = []
        for reg in self._bps_data:
            if not (
                pattern.search(reg.medicamento)
                or pattern.search(reg.principio_ativo)
            ):
                continue
            if apresentacao and apresentacao.lower() not in reg.apresentacao.lower():
                continue
            if uf and reg.uf.upper() != uf.upper():
                continue
            matches.append(reg)

        if not matches:
            return None

        precos = [r.preco_unitario for r in matches if r.preco_unitario > 0]
        if not precos:
            return None

        return BPSResumo(
            medicamento=medicamento,
            apresentacao=apresentacao or "",
            media=round(statistics.mean(precos), 4),
            mediana=round(statistics.median(precos), 4),
            minimo=round(min(precos), 4),
            maximo=round(max(precos), 4),
            desvio_padrao=round(
                statistics.stdev(precos) if len(precos) > 1 else 0.0,
                4,
            ),
            n_registros=len(precos),
            periodo=f"ultimos {periodo_meses} meses",
        )

    def get_registros_bps(
        self,
        medicamento: str,
        apresentacao: Optional[str] = None,
        limite: int = 20,
    ) -> list[BPSPreco]:
        """
        Busca registros individuais de compra nos dados BPS.

        Args:
            medicamento: Nome ou principio ativo
            apresentacao: Forma farmaceutica
            limite: Maximo de registros

        Returns:
            Lista de BPSPreco
        """
        pattern = re.compile(
            re.escape(medicamento), re.IGNORECASE
        )
        results: list[BPSPreco] = []
        for reg in self._bps_data:
            if not (
                pattern.search(reg.medicamento)
                or pattern.search(reg.principio_ativo)
            ):
                continue
            if apresentacao and apresentacao.lower() not in reg.apresentacao.lower():
                continue
            results.append(reg)
            if len(results) >= limite:
                break
        return results

    # ------------------------------------------------------------------
    # CMED methods
    # ------------------------------------------------------------------

    def load_cmed_csv(self, csv_path: str | Path):
        """
        Carrega tabela CMED de um arquivo CSV/XLS convertido.

        O CSV deve ter colunas: SUBSTANCIA, PRODUTO, APRESENTACAO,
        LABORATORIO, PF_SEM_IMPOSTOS, PMVG_SEM_IMPOSTOS,
        PMVG_COM_IMPOSTOS, LISTA_CONCESSAO.

        Args:
            csv_path: Caminho para o CSV
        """
        path = Path(csv_path)
        if not path.exists():
            logger.warning("CMED CSV not found: %s", path)
            return

        encoding = "utf-8"
        try:
            path.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            encoding = "latin-1"

        with path.open(encoding=encoding) as fh:
            reader = csv.DictReader(fh, delimiter=";")
            for row in reader:
                substancia = (
                    row.get("SUBSTANCIA", "")
                    or row.get("PRINCIPIO_ATIVO", "")
                ).strip()
                produto = (
                    row.get("PRODUTO", "")
                    or row.get("MEDICAMENTO", "")
                ).strip()
                apresentacao = (
                    row.get("APRESENTACAO", "")
                ).strip()

                def _parse_price(key):
                    v = row.get(key, "0").strip()
                    v = v.replace(".", "").replace(",", ".")
                    try:
                        return float(v)
                    except ValueError:
                        return 0.0

                entry = CMEDPreco(
                    medicamento=f"{substancia} - {produto}",
                    apresentacao=apresentacao,
                    laboratorio=row.get(
                        "LABORATORIO", ""
                    ).strip(),
                    pf_sem_impostos=_parse_price(
                        "PF_SEM_IMPOSTOS"
                    ),
                    pmvg_sem_impostos=_parse_price(
                        "PMVG_SEM_IMPOSTOS"
                    ),
                    pmvg_com_impostos=_parse_price(
                        "PMVG_COM_IMPOSTOS"
                    ),
                    lista_concessao=row.get(
                        "LISTA_CONCESSAO", ""
                    ).strip(),
                    data_publicacao=row.get(
                        "DATA_PUBLICACAO",
                        datetime.now().strftime("%Y-%m-%d"),
                    ).strip(),
                )

                key = self._cmed_key(
                    substancia or produto, apresentacao
                )
                self._cmed_data[key] = entry

        self._cmed_loaded = True
        logger.info(
            "Loaded %d CMED entries from %s",
            len(self._cmed_data), path,
        )

    async def ensure_cmed_loaded(self):
        """Load CMED data from local cache if available."""
        if self._cmed_loaded:
            return

        cache_dir = CMED_CACHE_DIR
        if cache_dir.exists():
            csvs = sorted(cache_dir.glob("*.csv"), reverse=True)
            if csvs:
                self.load_cmed_csv(csvs[0])
                return
            xls_files = sorted(
                cache_dir.glob("*.xls*"), reverse=True
            )
            if xls_files:
                self._load_cmed_xls(xls_files[0])

    def _load_cmed_xls(self, path: Path):
        """Parse CMED XLS using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            logger.warning(
                "openpyxl not installed; cannot parse %s", path
            )
            return

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        if ws is None:
            return

        headers = [
            str(c.value or "").upper().strip()
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]

        col_map = {}
        for i, h in enumerate(headers):
            if "SUBSTANCIA" in h or "PRINCIPIO" in h:
                col_map["substancia"] = i
            elif "PRODUTO" in h or "MEDICAMENTO" in h:
                col_map["produto"] = i
            elif "APRESENTA" in h:
                col_map["apresentacao"] = i
            elif "LABORAT" in h:
                col_map["laboratorio"] = i
            elif "PMVG" in h and "SEM" in h:
                col_map["pmvg_sem"] = i
            elif "PMVG" in h:
                col_map["pmvg_com"] = i
            elif "PF" in h and "SEM" in h:
                col_map["pf_sem"] = i

        for row in ws.iter_rows(min_row=2, values_only=True):
            substancia = str(
                row[col_map.get("substancia", 0)] or ""
            ).strip()
            produto = str(
                row[col_map.get("produto", 1)] or ""
            ).strip()
            apresentacao = str(
                row[col_map.get("apresentacao", 2)] or ""
            ).strip()

            def _val(col_name):
                idx = col_map.get(col_name)
                if idx is None:
                    return 0.0
                v = row[idx]
                try:
                    return float(v) if v else 0.0
                except (ValueError, TypeError):
                    return 0.0

            entry = CMEDPreco(
                medicamento=f"{substancia} - {produto}",
                apresentacao=apresentacao,
                laboratorio=str(
                    row[col_map.get("laboratorio", 3)] or ""
                ).strip(),
                pf_sem_impostos=_val("pf_sem"),
                pmvg_sem_impostos=_val("pmvg_sem"),
                pmvg_com_impostos=_val("pmvg_com"),
                lista_concessao="",
                data_publicacao=datetime.now().strftime(
                    "%Y-%m-%d"
                ),
            )
            key = self._cmed_key(
                substancia or produto, apresentacao
            )
            self._cmed_data[key] = entry

        self._cmed_loaded = True
        logger.info(
            "Loaded %d CMED entries from XLS %s",
            len(self._cmed_data), path,
        )

    def get_cmed_teto(
        self,
        medicamento: str,
        apresentacao: Optional[str] = None,
    ) -> Optional[CMEDPreco]:
        """
        Busca preco maximo CMED para medicamento.

        O PMVG (Preco Maximo de Venda ao Governo) e o teto
        obrigatorio para compras publicas de medicamentos.

        Args:
            medicamento: Nome ou principio ativo
            apresentacao: Forma farmaceutica

        Returns:
            CMEDPreco com teto ou None
        """
        # Try exact key match first
        key = self._cmed_key(medicamento, apresentacao or "")
        if key in self._cmed_data:
            return self._cmed_data[key]

        # Fuzzy search
        pattern = re.compile(
            re.escape(medicamento), re.IGNORECASE
        )
        for entry_key, entry in self._cmed_data.items():
            if pattern.search(entry.medicamento):
                if apresentacao:
                    if apresentacao.lower() in entry.apresentacao.lower():
                        return entry
                else:
                    return entry

        return None

    def verificar_teto(
        self,
        medicamento: str,
        preco_proposto: float,
        apresentacao: Optional[str] = None,
    ) -> dict:
        """
        Verifica se preco proposto esta dentro do teto CMED.

        Args:
            medicamento: Nome do medicamento
            preco_proposto: Preco a verificar
            apresentacao: Forma farmaceutica

        Returns:
            dict com resultado da verificacao
        """
        teto = self.get_cmed_teto(medicamento, apresentacao)
        if teto is None:
            return {
                "verificado": False,
                "erro": (
                    "Teto CMED nao encontrado para este "
                    "medicamento. Verifique se a tabela CMED "
                    "esta carregada em data/cmed/."
                ),
            }

        dentro = preco_proposto <= teto.pmvg_sem_impostos
        return {
            "verificado": True,
            "dentro_do_teto": dentro,
            "preco_proposto": preco_proposto,
            "teto_pmvg": teto.pmvg_sem_impostos,
            "diferenca": round(
                preco_proposto - teto.pmvg_sem_impostos, 4
            ),
            "data_referencia": teto.data_publicacao,
        }

    @staticmethod
    def _cmed_key(
        medicamento: str, apresentacao: str
    ) -> str:
        """Create a normalized lookup key."""
        return (
            f"{medicamento.strip().upper()}"
            f"|{apresentacao.strip().upper()}"
        )
