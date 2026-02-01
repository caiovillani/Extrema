"""
Cliente para consulta de precos SICRO.

O SICRO (Sistema de Custos Referenciais de Obras) e mantido pelo
DNIT (Departamento Nacional de Infraestrutura de Transportes).
Referencial de custos para infraestrutura de transportes.

Dados disponiveis em:
https://www.gov.br/dnit/pt-br/assuntos/planejamento-e-pesquisa/custos-e-pagamentos/
"""

import csv
import logging
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from tools.http_utils import CachedHTTPClient, HTTPError

logger = logging.getLogger(__name__)

# URL base para download das tabelas SICRO (DNIT)
SICRO_DOWNLOAD_BASE = (
    "https://www.gov.br/dnit/pt-br/assuntos/"
    "planejamento-e-pesquisa/custos-e-pagamentos"
)

# Diretorio para cache local de planilhas baixadas
SICRO_CACHE_DIR = Path(
    os.environ.get("SICRO_CACHE_DIR", "data/sicro")
)


@dataclass
class SICROComposicao:
    """Representa uma composicao SICRO."""

    codigo: str
    descricao: str
    unidade: str
    preco_unitario: float
    referencia_mes: str
    estado: str


class SICROClient:
    """Cliente para consultas na base SICRO.

    O fluxo e:
    1. Baixar planilha do DNIT (ou usar cache local)
    2. Parsear CSV/XLS para extrair composicoes
    3. Indexar em memoria para consultas rapidas
    """

    def __init__(
        self,
        estado: str = "MG",
        http: Optional[CachedHTTPClient] = None,
    ):
        self.estado = estado
        self.referencia_mes = datetime.now().strftime("%Y-%m")
        self.http = http or CachedHTTPClient()
        self._composicoes: dict[str, SICROComposicao] = {}
        self._loaded = False

    def load_from_csv(self, csv_path: str | Path):
        """
        Carrega composicoes a partir de um arquivo CSV local.

        O CSV deve ter colunas separadas por ponto-e-virgula com
        headers contendo CODIGO, DESCRICAO, UNIDADE, PRECO.

        Args:
            csv_path: Caminho para o arquivo CSV
        """
        path = Path(csv_path)
        if not path.exists():
            logger.warning("CSV not found: %s", path)
            return

        with path.open(encoding="latin-1") as fh:
            reader = csv.DictReader(fh, delimiter=";")
            for row in reader:
                codigo = (
                    row.get("CODIGO", "")
                    or row.get("CODIGO COMPOSICAO", "")
                    or row.get("CODIGO DA COMPOSICAO", "")
                ).strip()
                if not codigo:
                    continue

                descricao = (
                    row.get("DESCRICAO", "")
                    or row.get("DESCRICAO DA COMPOSICAO", "")
                    or row.get("DESCRICAO DO SERVICO", "")
                ).strip()

                unidade = (
                    row.get("UNIDADE", "")
                    or row.get("UNIDADE DE MEDIDA", "")
                ).strip()

                preco_str = (
                    row.get("PRECO UNITARIO", "")
                    or row.get("CUSTO TOTAL", "")
                    or row.get("PRECO", "")
                ).strip()
                preco_str = (
                    preco_str.replace(".", "").replace(",", ".")
                )
                try:
                    preco = float(preco_str) if preco_str else 0.0
                except ValueError:
                    preco = 0.0

                self._composicoes[codigo] = SICROComposicao(
                    codigo=codigo,
                    descricao=descricao,
                    unidade=unidade,
                    preco_unitario=preco,
                    referencia_mes=self.referencia_mes,
                    estado=self.estado,
                )

        self._loaded = True
        logger.info(
            "Loaded %d SICRO composicoes from %s",
            len(self._composicoes), path,
        )

    async def ensure_loaded(self):
        """
        Ensure data is loaded. Tries local cache first, then
        downloads from DNIT website.
        """
        if self._loaded:
            return

        # Try local cache directory
        cache_dir = SICRO_CACHE_DIR / self.estado.lower()
        if cache_dir.exists():
            csvs = sorted(cache_dir.glob("*.csv"), reverse=True)
            if csvs:
                self.load_from_csv(csvs[0])
                return
            xls_files = sorted(
                cache_dir.glob("*.xls*"), reverse=True
            )
            if xls_files:
                self._load_xls(xls_files[0])
                return

        # Try downloading
        await self._download_latest()

    async def _download_latest(self):
        """
        Attempt to download the latest SICRO table from DNIT.
        """
        estado = self.estado.upper()
        ref = datetime.now().strftime("%Y%m")
        filename = f"SICRO_Composicoes_{estado}_{ref}.csv"
        url = f"{SICRO_DOWNLOAD_BASE}/{filename}"

        try:
            raw = await self.http.get_bytes(url)
            cache_dir = SICRO_CACHE_DIR / estado.lower()
            cache_dir.mkdir(parents=True, exist_ok=True)
            out_path = cache_dir / filename
            out_path.write_bytes(raw)
            self.load_from_csv(out_path)
            logger.info(
                "Downloaded SICRO table: %s", out_path
            )
        except HTTPError as exc:
            logger.warning(
                "Could not download SICRO table (%s). "
                "Place CSV files in %s manually.",
                exc, SICRO_CACHE_DIR,
            )

    def _load_xls(self, path: Path):
        """Attempt to parse an XLS file using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            logger.warning(
                "openpyxl not installed; cannot parse %s", path
            )
            return

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        if ws is None:
            return

        headers = [
            str(c.value or "").upper().strip()
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        code_col = None
        desc_col = None
        unit_col = None
        price_col = None

        for i, h in enumerate(headers):
            if "CODIGO" in h:
                code_col = i
            elif "DESCRI" in h:
                desc_col = i
            elif "UNIDADE" in h or h == "UN":
                unit_col = i
            elif "PRECO" in h or "CUSTO" in h:
                price_col = i

        if code_col is None:
            return

        for row in ws.iter_rows(min_row=2, values_only=True):
            codigo = str(row[code_col] or "").strip()
            if not codigo:
                continue
            descricao = str(
                row[desc_col] if desc_col is not None else ""
            ).strip()
            unidade = str(
                row[unit_col] if unit_col is not None else ""
            ).strip()
            preco_raw = (
                row[price_col]
                if price_col is not None
                else 0
            )
            try:
                preco = float(preco_raw) if preco_raw else 0.0
            except (ValueError, TypeError):
                preco = 0.0

            self._composicoes[codigo] = SICROComposicao(
                codigo=codigo,
                descricao=descricao,
                unidade=unidade,
                preco_unitario=preco,
                referencia_mes=self.referencia_mes,
                estado=self.estado,
            )

        self._loaded = True
        logger.info(
            "Loaded %d SICRO composicoes from XLS %s",
            len(self._composicoes), path,
        )

    def get_composicao(
        self,
        codigo: str,
    ) -> Optional[SICROComposicao]:
        """
        Busca composicao por codigo SICRO.

        Args:
            codigo: Codigo SICRO (ex: "5914242")

        Returns:
            SICROComposicao ou None se nao encontrado
        """
        return self._composicoes.get(codigo)

    def search_composicoes(
        self,
        termo: str,
        limite: int = 10,
    ) -> list[SICROComposicao]:
        """
        Busca composicoes por descricao.

        Args:
            termo: Termo de busca na descricao
            limite: Maximo de resultados

        Returns:
            Lista de SICROComposicao
        """
        pattern = re.compile(re.escape(termo), re.IGNORECASE)
        results: list[SICROComposicao] = []
        for comp in self._composicoes.values():
            if pattern.search(comp.descricao):
                results.append(comp)
                if len(results) >= limite:
                    break
        return results
